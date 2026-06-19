#!/usr/bin/env python3
"""
Revisor automàtic d'entregues DWES (Symfony) — GitHub Classroom.

Llista els repositoris d'una organització de GitHub Classroom que comencen
per un prefix (el de l'assignment), els clona temporalment un a un, envia
el codi rellevant a l'API de Claude junt amb la rúbrica de RA, i genera un
Excel amb una fila per alumne.

Ús:
    export ANTHROPIC_API_KEY="sk-ant-..."
    export GITHUB_TOKEN="ghp_..."
    python revisar_entregas.py --org la-meua-org --prefix whatsapp-clone --rubrica rubrica.md --salida informe_revisio.xlsx
"""
import argparse
import csv
import json
import os
import subprocess
import sys
import tempfile
import time
from datetime import date
from pathlib import Path

import anthropic
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

MODEL = "claude-sonnet-4-6"
MAX_TOTAL_CHARS = 180_000
EXCLUDE_DIRS = {"vendor", "var", "node_modules", ".git", "build", "bundles"}
INCLUDE_EXT = {".php", ".twig", ".yaml", ".yml", ".json", ".sql", ".env.example"}
PRIORITY_DIRS = ["src", "templates", "config", "migrations"]
RA_IDS = [f"RA{i}" for i in range(1, 10)]
ESTAT_COLOR = {"Assolit": "C6EFCE", "Parcial": "FFEB9C", "No assolit": "FFC7CE"}
GITHUB_API = "https://api.github.com"


# ---------- GitHub Classroom ----------

def llistar_repos(org: str, prefix: str, token: str) -> list:
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/vnd.github+json"}
    repos, page = [], 1
    while True:
        resp = requests.get(
            f"{GITHUB_API}/orgs/{org}/repos",
            headers=headers,
            params={"per_page": 100, "page": page, "type": "all"},
            timeout=30,
        )
        resp.raise_for_status()
        dades = resp.json()
        if not dades:
            break
        repos += [r for r in dades if r["name"].startswith(prefix)]
        page += 1
    return repos


def clonar_repo(clone_url: str, token: str, dest: Path) -> None:
    url_auth = clone_url.replace("https://", f"https://x-access-token:{token}@")
    subprocess.run(
        ["git", "clone", "--depth", "1", "--quiet", url_auth, str(dest)],
        check=True, capture_output=True, text=True,
    )


def carregar_roster(path: str) -> dict:
    if not path:
        return {}
    with open(path, newline="", encoding="utf-8") as f:
        return {row["github_username"]: row["nom_alumne"] for row in csv.DictReader(f)}


def nom_alumne(repo_name: str, prefix: str, roster: dict) -> str:
    identificador = repo_name[len(prefix):].lstrip("-_")
    return roster.get(identificador, identificador)


# ---------- Recollida de codi ----------

def hauria_dexcloure(rel_path: Path) -> bool:
    return any(part in EXCLUDE_DIRS for part in rel_path.parts)


def recollir_codi(carpeta: Path) -> str:
    candidats = []
    for p in carpeta.rglob("*"):
        if not p.is_file():
            continue
        rel = p.relative_to(carpeta)
        if hauria_dexcloure(rel) or p.suffix not in INCLUDE_EXT:
            continue
        prioritat = next(
            (i for i, d in enumerate(PRIORITY_DIRS) if rel.parts and rel.parts[0] == d),
            len(PRIORITY_DIRS),
        )
        candidats.append((prioritat, rel, p))
    candidats.sort(key=lambda x: x[0])

    bloc, total = [], 0
    for _, rel, p in candidats:
        try:
            contingut = p.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            continue
        fragment = f"\n--- {rel} ---\n{contingut}\n"
        if total + len(fragment) > MAX_TOTAL_CHARS:
            break
        bloc.append(fragment)
        total += len(fragment)
    return "".join(bloc)


# ---------- Avaluació amb Claude ----------

def construir_sistema(rubrica: str) -> str:
    return f"""Ets un professor de DAW que avalua entregues del mòdul DWES (0613, RD 405/2023).
Avalua el codi Symfony de l'alumne segons aquesta rúbrica de Resultats d'Aprenentatge:

{rubrica}

Respon NOMÉS amb un JSON vàlid (sense ``` ni text addicional) amb aquesta estructura:
{{
  "resultats": [
    {{"ra": "RA1", "estat": "Assolit|Parcial|No assolit", "evidencia": "frase curta amb referència a fitxer/funció", "suggeriment": "frase curta"}}
  ],
  "comentari_general": "2-3 frases amb punts forts i febles globals"
}}
Inclou un objecte per a cada RA1..RA9. Si el projecte no aborda un RA, marca'l "No assolit" i indica-ho a l'evidència."""


def avaluar_alumne(client: anthropic.Anthropic, sistema: str, nom: str, codi: str, intents: int = 3) -> dict:
    ultim_error = None
    for intent in range(intents):
        try:
            resposta = client.messages.create(
                model=MODEL,
                max_tokens=3000,
                system=sistema,
                messages=[{"role": "user", "content": f"Alumne: {nom}\n\nCodi del projecte:\n{codi or '(cap fitxer rellevant trobat)'}"}],
            )
            text = "".join(b.text for b in resposta.content if b.type == "text").strip()
            text = text.removeprefix("```json").removeprefix("```").removesuffix("```").strip()
            return json.loads(text)
        except (json.JSONDecodeError, anthropic.APIStatusError, anthropic.APIConnectionError) as e:
            ultim_error = e
            time.sleep(2 ** intent)
    raise RuntimeError(f"Fallada després de {intents} intents: {ultim_error}")


# ---------- Excel ----------

def generar_excel(files: list, sortida: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Revisió DWES"

    capçalera = ["Alumne"]
    for ra in RA_IDS:
        capçalera += [f"{ra} estat", f"{ra} evidència", f"{ra} suggeriment"]
    capçalera += ["Comentari general", "Data revisió", "Error"]
    ws.append(capçalera)

    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for fila in files:
        ws.append(fila["valors"])

    for row_idx, fila in enumerate(files, start=2):
        for col_idx, ra in enumerate(RA_IDS):
            estat = fila["estats"].get(ra)
            col = 2 + col_idx * 3
            if estat in ESTAT_COLOR:
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", start_color=ESTAT_COLOR[estat])

    ws.column_dimensions["A"].width = 22
    for i in range(2, len(capçalera) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 28
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(sortida)


# ---------- Main ----------

def main():
    parser = argparse.ArgumentParser(description="Revisor automàtic d'entregues DWES (GitHub Classroom)")
    parser.add_argument("--org", required=True, help="Organització de GitHub Classroom")
    parser.add_argument("--prefix", required=True, help="Prefix del nom dels repos (el de l'assignment)")
    parser.add_argument("--rubrica", default="rubrica.md")
    parser.add_argument("--salida", default="informe_revisio.xlsx")
    parser.add_argument("--roster", default=None, help="CSV opcional amb columnes github_username,nom_alumne")
    args = parser.parse_args()

    github_token = os.environ.get("GITHUB_TOKEN")
    if not github_token:
        sys.exit("Defineix GITHUB_TOKEN amb un token d'accés (scope repo) per a llegir els repositoris de l'organització.")

    rubrica = Path(args.rubrica).read_text(encoding="utf-8")
    sistema = construir_sistema(rubrica)
    client = anthropic.Anthropic()
    roster = carregar_roster(args.roster)

    print(f"Buscant repos a '{args.org}' amb prefix '{args.prefix}'...")
    repos = llistar_repos(args.org, args.prefix, github_token)
    if not repos:
        sys.exit(f"Cap repositori trobat a '{args.org}' amb el prefix '{args.prefix}'.")
    print(f"Trobats {len(repos)} repositoris.\n")

    files_resultat = []
    for repo in repos:
        nom = nom_alumne(repo["name"], args.prefix, roster)
        print(f"Revisant {nom} ({repo['name']})...")
        with tempfile.TemporaryDirectory() as tmp:
            dest = Path(tmp) / "repo"
            try:
                clonar_repo(repo["clone_url"], github_token, dest)
                codi = recollir_codi(dest)
            except subprocess.CalledProcessError as e:
                err = e.stderr.strip() if e.stderr else str(e)
                print(f"  [error clonant] {err}")
                valors = [nom] + ["", "", ""] * len(RA_IDS) + ["", str(date.today()), f"Error clonant: {err}"]
                files_resultat.append({"valors": valors, "estats": {}})
                continue

            try:
                dades = avaluar_alumne(client, sistema, nom, codi)
                estats = {r["ra"]: r["estat"] for r in dades["resultats"]}
                valors = [nom]
                for r in dades["resultats"]:
                    valors += [r["estat"], r["evidencia"], r["suggeriment"]]
                valors += [dades.get("comentari_general", ""), str(date.today()), ""]
                files_resultat.append({"valors": valors, "estats": estats})
            except Exception as e:
                print(f"  [error avaluant] {e}")
                valors = [nom] + ["", "", ""] * len(RA_IDS) + ["", str(date.today()), str(e)]
                files_resultat.append({"valors": valors, "estats": {}})

    generar_excel(files_resultat, Path(args.salida))
    print(f"\nFet. Informe guardat a {args.salida}")


if __name__ == "__main__":
    main()
